# -*- coding: utf-8 -*-
"""
SBFlashPro.py (stable build v0.15)

v0.15 変更点
- 上段ボタン列に「ランダム出題」追加
- ランダム出題はトグル式（ON/OFF表示）
- ランダムON時も「この論点だけ復習」「絞り込み解除」と共存
- シート切替後もランダムモード状態を維持して並びを再構成

v0.08 変更点
- 上段（設問）にスクロールバー追加
- 上段（設問）と下段（正解/解説）を同じ高さに（gridで等分）
- 中段（解答入力）は4行のまま
- F3 を「回答保存(F3)」に変更（保存はF1判定後のみ）
- 回答シート（wrong_sheet）は ini の wrong_sheet 名をそのまま使う
- 回答シートに consecutive_ok（連続正解回数）列を追加して書き込み
  - 正解保存: consecutive_ok += 1 / last_ok 更新
  - 不正解保存: mistakes += 1 / consecutive_ok = 0 / last_miss 更新
- 削除はExcel手動（削除ボタンなし運用に合わせる）
"""

import argparse
import csv
import os
import random
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
import sys
import time
import configparser

import pandas as pd
from openpyxl import load_workbook, Workbook

import tkinter as tk
from tkinter import messagebox, ttk

from PIL import Image, ImageTk


DEFAULT_INI = "SBFlashPro.ini"

# =====================================
# SBFlash Pro Version (on-code)
# =====================================
APP_VERSION = "0.15"

# =====================================
# SBKnowledgeData Layout (0 origin)
# =====================================
COL_QUESTION_NO = 0
COL_QUESTION = 1
COL_ANSWER = 2
COL_QUESTION_IMAGE = 3
COL_ANSWER_IMAGE = 4
COL_EXPLANATION = 5
COL_MNEMONIC = 6
COL_SUBJECT = 7
COL_TAGS = 8
COL_KEYWORDS = 9
MIN_CARD_COLUMNS = 10

# ===== v0.4 追加：データ開始行（Excel上の行番号） =====
DATA_START_ROW_DEFAULT = 2   # 通常シート：2行目からデータ
WRONG_START_ROW = 3          # 回答シート：3行目からデータ（2行目は日本語タイトル行想定）
# ========================================================


def _base_dir() -> Path:
    # .py実行時はこのpyの場所、exe化してもexeの場所になる
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def _create_default_ini(path: Path) -> None:
    # exe/pyをダブルクリックしただけで使える“最低限の初期値”
    text = (
        "[app]\n"
        "app_title=暗記カード\n"
        "EXCEL_PATH=FlashCards.xlsx\n"
        "initial_sheet=sheet0\n"
        "wrong_sheet=回答シート\n"
        "data_start_row_default=2\n"
        "wrong_start_row=3\n"
        "wrong_only=false\n"
        "worst_first=false\n"
        "all_subjects=false\n"
        "\n"
        "[ui]\n"
        "auto_ratio=0.90\n"
        "min_width=820\n"
        "min_height=620\n"
        "max_width=2200\n"
        "max_height=1400\n"
        "thumb_size=400\n"
        "zoom_max=1200\n"
    )
    path.write_text(text, encoding="utf-8")


def load_settings() -> dict:
    ini_path = _base_dir() / DEFAULT_INI
    if not ini_path.exists():
        _create_default_ini(ini_path)

    cfg = configparser.ConfigParser()
    cfg.read(ini_path, encoding="utf-8")

    if "app" not in cfg:
        _create_default_ini(ini_path)
        cfg.read(ini_path, encoding="utf-8")

    app = cfg["app"]
    ui = cfg["ui"] if "ui" in cfg else None

    def get_bool(section, key: str, default: bool) -> bool:
        try:
            if section is None:
                return default
            return section.getboolean(key, fallback=default)
        except Exception:
            return default

    def get_int(section, key: str, default: int) -> int:
        try:
            if section is None:
                return default
            v = str(section.get(key, fallback=str(default))).strip()
            if v == "":
                return default
            return int(v)
        except Exception:
            return default

    def get_float(section, key: str, default: float) -> float:
        try:
            if section is None:
                return default
            v = str(section.get(key, fallback=str(default))).strip()
            if v == "":
                return default
            return float(v)
        except Exception:
            return default

    # excel_path / EXCEL_PATH を受け付ける。相対パスは ini → exe/py → 親フォルダ の順で解決する
    excel_path_raw = app.get("excel_path", fallback=app.get("EXCEL_PATH", "FlashCards.xlsx")).strip()
    excel_path = Path(excel_path_raw)
    if not excel_path.is_absolute():
        candidates = [
            ini_path.parent / excel_path,
            _base_dir() / excel_path,
            _base_dir().parent / excel_path,
            Path.cwd() / excel_path,
        ]
        resolved = None
        for c in candidates:
            if c.exists():
                resolved = c
                break
        excel_path = resolved if resolved is not None else candidates[0]

    ui_settings = {
        "window_width": get_int(ui, "window_width", 0),
        "window_height": get_int(ui, "window_height", 0),
        "start_maximized": get_bool(ui, "start_maximized", False),
        "auto_ratio": get_float(ui, "auto_ratio", 0.90),
        "min_width": get_int(ui, "min_width", 820),
        "min_height": get_int(ui, "min_height", 620),
        "max_width": get_int(ui, "max_width", 2200),
        "max_height": get_int(ui, "max_height", 1400),
        "thumb_size": get_int(ui, "thumb_size", 400),
        "zoom_max": get_int(ui, "zoom_max", 1200),
    }

    return {
        "ini_path": str(ini_path),
        "app_title": app.get("app_title", "暗記カード").strip(),
        "app_version": APP_VERSION,
        "EXCEL_PATH": str(excel_path),
        "initial_sheet": app.get("initial_sheet", "sheet0").strip(),
        "wrong_sheet": app.get("wrong_sheet", "回答シート").strip(),
        "data_start_row_default": get_int(app, "data_start_row_default", DATA_START_ROW_DEFAULT),
        "wrong_start_row": get_int(app, "wrong_start_row", WRONG_START_ROW),
        "wrong_only": get_bool(app, "wrong_only", False),
        "worst_first": get_bool(app, "worst_first", False),
        "all_subjects": get_bool(app, "all_subjects", False),
        "ui": ui_settings,
    }


def parse_list_cell(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return []
    s = str(v).strip()
    if not s:
        return []
    s = s.replace("、", ",")
    return [p.strip() for p in s.split(",") if p.strip()]


def parse_tags(v):
    return parse_list_cell(v)


def parse_keywords(v):
    return parse_list_cell(v)


def normalize_answer(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.strip()

    # ○×問題は記号自体が解答になるため、先に代表表記へそろえる
    if s in ("○", "〇"):
        return "○"
    if s in ("×", "✕", "✖", "x", "X"):
        return "×"

    s = s.lower()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9a-z\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fff○×]+", "", s)
    return s


def sheet_key(s: str) -> str:
    """比較用。空白/全角空白を潰して小文字化。"""
    if s is None:
        return ""
    t = str(s)
    t = t.replace("\u3000", " ")
    t = re.sub(r"\s+", "", t)  # 全空白削除
    return t.strip().lower()


def list_question_sheets(excel_path: str) -> list[str]:
    """問題シート候補: 回答シート（wrong_sheet）以外、かつ先頭行に question がありそうなシートを優先したいが、
    ここでは単純に全シートを返し、UI操作で選ぶ前提にする（安定優先）。
    """
    p = Path(excel_path)
    if not p.exists():
        return []
    try:
        xls = pd.ExcelFile(p)
        return list(xls.sheet_names)
    except Exception:
        return []


def resolve_sheet_name(excel_path: str, sheet_arg: str | None) -> str:
    """sheet指定を実シート名に解決する。

    - None / '' / 'sheet0' / '0' : 先頭シート
    - 数字文字列              : 0-based index のシート
    - それ以外                : シート名として扱う
    """
    p = Path(excel_path)
    if not p.exists():
        return str(sheet_arg) if sheet_arg is not None else "sheet0"

    try:
        xls = pd.ExcelFile(p)
        names = xls.sheet_names
        if not names:
            return str(sheet_arg) if sheet_arg is not None else ""

        if sheet_arg is None:
            return names[0]

        s = str(sheet_arg).strip()
        if s == "" or s.lower() == "sheet0" or s == "0":
            return names[0]

        if s.isdigit():
            i = int(s)
            if 0 <= i < len(names):
                return names[i]
            raise ValueError(f"sheet index out of range: {i} (0..{len(names)-1})")

        if s in names:
            return s

        raise ValueError(f"Worksheet named '{s}' not found")
    except Exception:
        return str(sheet_arg) if sheet_arg is not None else "sheet0"


def _safe_iloc(row, col_index, default=""):
    try:
        value = row.iloc[col_index]
    except Exception:
        return default
    return default if pd.isna(value) else value


def has_trim_value(v) -> bool:
    return str(v if v is not None else "").strip() != ""


def extract_question_row(row, row_number_for_fallback: int):
    q_no = str(_safe_iloc(row, COL_QUESTION_NO, "")).strip()
    q = str(_safe_iloc(row, COL_QUESTION, "")).strip()
    a = str(_safe_iloc(row, COL_ANSWER, "")).strip()
    question_img_path = str(_safe_iloc(row, COL_QUESTION_IMAGE, "")).strip()
    answer_img_path = str(_safe_iloc(row, COL_ANSWER_IMAGE, "")).strip()
    explanation = str(_safe_iloc(row, COL_EXPLANATION, "")).strip()
    mnemonic = str(_safe_iloc(row, COL_MNEMONIC, "")).strip()
    subj = str(_safe_iloc(row, COL_SUBJECT, "")).strip()
    tags_value = _safe_iloc(row, COL_TAGS, None)
    keywords_value = _safe_iloc(row, COL_KEYWORDS, None)
    tags = [] if tags_value is None else parse_tags(tags_value)
    keywords = [] if keywords_value is None else parse_keywords(keywords_value)

    if not q_no:
        q_no = str(row_number_for_fallback)

    return {
        "question_no": q_no,
        "question": q,
        "answer": a,
        "question_image_path": question_img_path,
        "answer_image_path": answer_img_path,
        "image_path": answer_img_path,
        "explanation": explanation,
        "mnemonic": mnemonic,
        "subject": subj,
        "tags": tags,
        "keywords": keywords,
    }


def load_cards(excel_path: str, sheet_name: str, data_start_row: int = DATA_START_ROW_DEFAULT):
    """
    出題シート（1行目ヘッダー）
      A: question_no（空なら行番号で自動採番）
      B: question
      C: answer
      D: question_image_path(optional)  ※設問画像ファイルパス（相対はexe/py基準）
      E: answer_image_path(optional)    ※解答画像ファイルパス（相対はexe/py基準）
      F: explanation(optional)          ※解説
      G: mnemonic(optional)             ※語呂合せ
      H: subject(optional)
      I: tags(optional)
      J: keywords(optional)
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {path.resolve()}")

    df = pd.read_excel(path, sheet_name=sheet_name)
    # データ開始行を調整（1行目はヘッダー扱い）
    df = df.iloc[max(0, data_start_row - 2):].reset_index(drop=True)
    while df.shape[1] < MIN_CARD_COLUMNS:
        df[df.shape[1]] = None
    df = df.dropna(how="all")

    cards = []
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        card = extract_question_row(row, i)
        if not card["question"]:
            continue

        card["source_sheet"] = str(sheet_name)
        cards.append(card)

    return cards


# v0.08: consecutive_ok 追加
WRONG_COLUMNS = [
    "source_sheet",
    "question_no",
    "question",
    "answer",
    "question_image_path",
    "answer_image_path",
    "image_path",
    "explanation",
    "mnemonic",
    "subject",
    "tags",
    "keywords",
    "mistakes",
    "consecutive_ok",
    "last_miss",
    "last_ok",
]


def read_wrong_all(excel_path: str, wrong_sheet_name: str, wrong_start_row: int = WRONG_START_ROW) -> pd.DataFrame:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {path.resolve()}")

    try:
        df = pd.read_excel(path, sheet_name=wrong_sheet_name)
        df = df.iloc[max(0, wrong_start_row - 2):].reset_index(drop=True)
    except Exception:
        df = pd.DataFrame(columns=WRONG_COLUMNS)

    # 旧列名互換
    if "question_no" not in df.columns and "no" in df.columns:
        df = df.rename(columns={"no": "question_no"})

    if "answer_image_path" not in df.columns and "image_path" in df.columns:
        df["answer_image_path"] = df["image_path"]
    if "question_image_path" not in df.columns:
        df["question_image_path"] = None

    for col in WRONG_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[WRONG_COLUMNS].copy()

    df["source_sheet"] = df["source_sheet"].fillna("").astype(str).str.strip()
    df["question_no"] = df["question_no"].fillna("").astype(str).str.strip()

    # 数値列の補正
    def _to_int(x, default=0):
        try:
            if pd.isna(x) or x is None or str(x).strip() == "":
                return default
            return int(float(x))
        except Exception:
            return default

    df["mistakes"] = df["mistakes"].apply(lambda v: _to_int(v, 0))
    df["consecutive_ok"] = df["consecutive_ok"].apply(lambda v: _to_int(v, 0))

    return df


def write_wrong_sheet(
    excel_path: str,
    wrong_sheet_name: str,
    df: pd.DataFrame,
    *,
    title_row: int = 2,
    data_start_row: int = WRONG_START_ROW,
) -> None:
    """回答シートを書き込む（列幅・書式を壊さない版）。値だけ更新する。"""
    excel_path = str(excel_path)
    path = Path(excel_path)

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    if wrong_sheet_name in wb.sheetnames:
        ws = wb[wrong_sheet_name]
    else:
        ws = wb.create_sheet(title=wrong_sheet_name)

    # 2行目（日本語タイトル行）を保持
    existing_title = [None] * len(WRONG_COLUMNS)
    header_map = {}
    try:
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            if isinstance(v, str) and v.strip():
                header_map[v.strip()] = col
        for j, col_name in enumerate(WRONG_COLUMNS, start=1):
            col_idx = header_map.get(col_name, j)
            existing_title[j - 1] = ws.cell(row=title_row, column=col_idx).value
    except Exception:
        pass

    # 1行目: ヘッダー
    for j, col_name in enumerate(WRONG_COLUMNS, start=1):
        ws.cell(row=1, column=j, value=col_name)

    # 2行目: 日本語タイトル（既存があれば維持）
    for j, _ in enumerate(WRONG_COLUMNS, start=1):
        ws.cell(row=title_row, column=j, value=existing_title[j - 1])

    # 既存データ部分を値だけクリア
    max_row = ws.max_row
    max_col = max(ws.max_column, len(WRONG_COLUMNS))
    for r in range(data_start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

    df2 = df.copy()
    for c in WRONG_COLUMNS:
        if c not in df2.columns:
            df2[c] = None
    df2 = df2[WRONG_COLUMNS]

    r = data_start_row
    for rec in df2.itertuples(index=False, name=None):
        for j, v in enumerate(rec, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1

    # 既定の空シートが残っていたら削除
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        sh = wb["Sheet"]
        if sh.max_row == 1 and sh.max_column == 1 and sh["A1"].value is None:
            wb.remove(sh)

    wb.save(path)


def upsert_answer_log(
    excel_path: str,
    wrong_sheet_name: str,
    *,
    source_sheet: str,
    question_no: str,
    q: str,
    a: str,
    question_image_path: str = "",
    answer_image_path: str = "",
    image_path: str = "",
    explanation: str = "",
    mnemonic: str = "",
    subject: str = "",
    tags_list=None,
    keywords_list=None,
    is_ok: bool,
    wrong_start_row: int = WRONG_START_ROW,
):
    """
    主キー: (source_sheet, question_no)
    - 正解保存: consecutive_ok += 1, last_ok 更新（mistakesは増やさない）
    - 不正解保存: mistakes += 1, consecutive_ok = 0, last_miss 更新
    """
    df = read_wrong_all(excel_path, wrong_sheet_name, wrong_start_row=wrong_start_row)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    src = str(source_sheet).strip()
    qno = str(question_no).strip()
    if not src:
        raise ValueError("source_sheet が空です（保存できません）")
    if not qno:
        raise ValueError("question_no が空です（保存できません）")

    tags_str = ",".join(tags_list or [])
    kw_str = ",".join(keywords_list or [])

    hit = (df["source_sheet"].apply(sheet_key) == sheet_key(src)) & (df["question_no"] == qno)

    if hit.any():
        idx = df.index[hit][0]

        # 共通上書き（最新の内容を保持）
        df.at[idx, "source_sheet"] = src
        df.at[idx, "question_no"] = qno
        df.at[idx, "question"] = q
        df.at[idx, "answer"] = a
        df.at[idx, "question_image_path"] = question_image_path
        df.at[idx, "answer_image_path"] = answer_image_path
        df.at[idx, "image_path"] = answer_image_path or image_path
        df.at[idx, "explanation"] = explanation
        df.at[idx, "mnemonic"] = mnemonic
        df.at[idx, "subject"] = subject
        df.at[idx, "tags"] = tags_str
        df.at[idx, "keywords"] = kw_str

        cur_m = int(df.at[idx, "mistakes"]) if str(df.at[idx, "mistakes"]).strip() != "" else 0
        cur_ok = int(df.at[idx, "consecutive_ok"]) if str(df.at[idx, "consecutive_ok"]).strip() != "" else 0

        if is_ok:
            df.at[idx, "consecutive_ok"] = cur_ok + 1
            df.at[idx, "last_ok"] = now
        else:
            df.at[idx, "mistakes"] = cur_m + 1
            df.at[idx, "consecutive_ok"] = 0
            df.at[idx, "last_miss"] = now

    else:
        df = pd.concat([df, pd.DataFrame([{
            "source_sheet": src,
            "question_no": qno,
            "question": q,
            "answer": a,
            "question_image_path": question_image_path,
            "answer_image_path": answer_image_path,
            "image_path": answer_image_path or image_path,
            "explanation": explanation,
            "mnemonic": mnemonic,
            "subject": subject,
            "tags": tags_str,
            "keywords": kw_str,
            "mistakes": (0 if is_ok else 1),
            "consecutive_ok": (1 if is_ok else 0),
            "last_miss": (now if not is_ok else ""),
            "last_ok": (now if is_ok else ""),
        }])], ignore_index=True)

    write_wrong_sheet(excel_path, wrong_sheet_name, df, data_start_row=wrong_start_row)


# ==================================================
# v0.05: ウィンドウサイズ（ini優先→無ければ画面から自動）+ 前回サイズ保存
# ==================================================
def _clamp(v: int, vmin: int, vmax: int) -> int:
    return max(vmin, min(vmax, v))


def apply_initial_geometry(root: tk.Tk, ui_settings: dict):
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()

    auto_ratio = float(ui_settings.get("auto_ratio", 0.90))
    min_w = int(ui_settings.get("min_width", 820))
    min_h = int(ui_settings.get("min_height", 620))
    max_w = int(ui_settings.get("max_width", 2200))
    max_h = int(ui_settings.get("max_height", 1400))

    w_ini = int(ui_settings.get("window_width", 0) or 0)
    h_ini = int(ui_settings.get("window_height", 0) or 0)

    if w_ini > 0 and h_ini > 0:
        w, h = w_ini, h_ini
    else:
        w, h = int(sw * auto_ratio), int(sh * auto_ratio)

    w = _clamp(w, min_w, max_w)
    h = _clamp(h, min_h, max_h)

    x = max(0, (sw - w) // 2)
    y = max(0, (sh - h) // 2)
    root.geometry(f"{w}x{h}+{x}+{y}")

    try:
        root.minsize(min_w, min_h)
    except Exception:
        pass

    if bool(ui_settings.get("start_maximized", False)):
        try:
            root.state("zoomed")
        except Exception:
            pass


def attach_geometry_saver(root: tk.Tk, ini_path: str, debounce_ms: int = 400):
    job = {"id": None}

    def save_now():
        try:
            cfg = configparser.ConfigParser()
            cfg.read(ini_path, encoding="utf-8")
            if "ui" not in cfg:
                cfg["ui"] = {}

            if str(root.state()).lower() == "zoomed":
                cfg["ui"]["start_maximized"] = "true"
            else:
                cfg["ui"]["start_maximized"] = "false"
                geo = root.geometry().split("+")[0]  # WxH
                w_str, h_str = geo.split("x")
                cfg["ui"]["window_width"] = str(int(w_str))
                cfg["ui"]["window_height"] = str(int(h_str))

            with open(ini_path, "w", encoding="utf-8") as f:
                cfg.write(f)
        except Exception:
            pass

    def schedule(_evt=None):
        if job["id"] is not None:
            try:
                root.after_cancel(job["id"])
            except Exception:
                pass
        job["id"] = root.after(debounce_ms, save_now)

    root.bind("<Configure>", schedule)

    def on_close():
        save_now()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)


# ==================================================
# v0.06: 画像パネル（D列=画像ファイルパス）
#  - 幅固定（サムネはその枠に収まるよう縮小）
#  - ボタンは最下部
#  - 回答表示時のみ show() される想定
# ==================================================
class ImagePanel:
    """右側に表示するサムネイル画像パネル（回答表示時のみ表示）。
    - 表示枠は「固定幅・固定高」
    - サムネは枠内に必ず収まるよう縮小（縦横比維持）
    - 「画像拡大」ボタンは常に表示（画像無し時は disabled）
    """

    def __init__(self, parent, *, thumb_size: int = 240, zoom_max: int = 1200, panel_width: int = 420):
        self.thumb_size = int(thumb_size)      # サムネ枠の高さ（px）
        self.zoom_max = int(zoom_max)
        self.panel_width = int(panel_width)

        # pack_propagate(False) を使う場合、height を明示しないと潰れることがあるため固定高を与える
        # ボタン領域分を加算
        self._button_h = 44
        self._pad_h = 28
        self.frame = tk.LabelFrame(parent, text="画像", width=self.panel_width, height=self.thumb_size + self._button_h + self._pad_h)
        self.frame.grid_propagate(False)
        self.frame.pack_propagate(False)
        self._visible = False

        # --- 内部は grid で固定（packと混ぜない） ---
        self.frame.grid_rowconfigure(0, weight=1)
        self.frame.grid_rowconfigure(1, weight=0)
        self.frame.grid_columnconfigure(0, weight=1)

        # 画像表示エリア（固定サイズ）
        self.image_box = tk.Frame(self.frame, width=self.panel_width - 16, height=self.thumb_size)
        self.image_box.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 4))
        self.image_box.grid_propagate(False)
        self.image_box.pack_propagate(False)

        self.thumb_label = tk.Label(self.image_box, text="(no image)", anchor="center")
        self.thumb_label.pack(fill="both", expand=True)

        # ボタン（常に表示、画像無し時は disabled）
        self.btn_zoom = tk.Button(self.frame, text="画像拡大", command=self.open_zoom, state="disabled")
        self.btn_zoom.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))

        self.current_path = ""
        self.thumb_imgtk = None
        self._cache = {}  # abs_path -> PIL.Image

    def grid(self, **kwargs):
        self._grid_kwargs = kwargs

    def show(self):
        if self._visible:
            return
        try:
            if hasattr(self, "_grid_kwargs"):
                self.frame.grid(**self._grid_kwargs)
            else:
                self.frame.pack(side="right", fill="y", padx=(8, 0), pady=0)
        except Exception:
            self.frame.pack(side="right", fill="y", padx=(8, 0), pady=0)
        self._visible = True

    def hide(self):
        if not self._visible:
            return
        try:
            if hasattr(self, "_grid_kwargs"):
                self.frame.grid_remove()
            else:
                self.frame.pack_forget()
        except Exception:
            pass
        self._visible = False

    def _resolve_path(self, path: str) -> str:
        p = str(path or "").strip()
        if not p:
            return ""
        pp = Path(p)
        if pp.is_absolute():
            return str(pp)
        return str(_base_dir() / pp)

    def _thumb_box_size(self) -> tuple[int, int]:
        """実表示枠サイズを返す（初回は指定値を返す）。"""
        try:
            self.frame.update_idletasks()
            w = int(self.image_box.winfo_width())
            h = int(self.image_box.winfo_height())
            if w > 10 and h > 10:
                return w, h
        except Exception:
            pass
        return max(80, self.panel_width - 16), max(80, self.thumb_size)

    def set_image(self, path: str):
        self.current_path = str(path or "").strip()

        if not self.current_path:
            self.thumb_label.config(image="", text="(no image)")
            self.thumb_imgtk = None
            try:
                self.btn_zoom.configure(state="disabled")
            except Exception:
                pass
            return

        abs_path = self._resolve_path(self.current_path)
        if not abs_path or not Path(abs_path).exists():
            self.thumb_label.config(image="", text="画像なし")
            self.thumb_imgtk = None
            try:
                self.btn_zoom.configure(state="disabled")
            except Exception:
                pass
            return

        img = self._cache.get(abs_path)
        if img is None:
            img = Image.open(abs_path)
            self._cache[abs_path] = img

        # サムネは「実枠」に必ず収まるよう縮小（縦横比維持）
        box_w, box_h = self._thumb_box_size()
        max_w = max(80, int(box_w - 6))
        max_h = max(80, int(box_h - 6))

        thumb = img.copy()
        thumb.thumbnail((max_w, max_h))
        self.thumb_imgtk = ImageTk.PhotoImage(thumb)
        self.thumb_label.config(image=self.thumb_imgtk, text="")
        self.thumb_label.image = self.thumb_imgtk

        try:
            self.btn_zoom.configure(state="normal")
        except Exception:
            pass

    def open_zoom(self):
        if not self.current_path:
            return
        abs_path = self._resolve_path(self.current_path)
        if not abs_path or not Path(abs_path).exists():
            return

        img = self._cache.get(abs_path)
        if img is None:
            img = Image.open(abs_path)
            self._cache[abs_path] = img

        top = tk.Toplevel()
        top.title("画像拡大")

        view = img.copy()
        view.thumbnail((self.zoom_max, self.zoom_max))
        imgtk = ImageTk.PhotoImage(view)

        lbl = tk.Label(top, image=imgtk)
        lbl.image = imgtk
        lbl.pack(padx=10, pady=10)


class FlashcardsApp(tk.Tk):
    def __init__(self, cards, *, excel_path: str, source_sheet: str, wrong_sheet: str, base_cards,
                 app_title: str = "暗記カード",
                 app_version: str = "",
                 ui_settings: dict | None = None,
                 ini_path: str | None = None,
                 data_start_row_default: int = DATA_START_ROW_DEFAULT,
                 wrong_start_row: int = WRONG_START_ROW,
                 initial_random_mode: bool = False):
        super().__init__()
        self.excel_path = excel_path
        self.source_sheet = str(source_sheet)
        self.wrong_sheet = str(wrong_sheet)
        self.app_title = str(app_title or "暗記カード").strip()
        self.app_version = str(app_version or "").strip()

        self.data_start_row_default = int(data_start_row_default)
        self.wrong_start_row = int(wrong_start_row)

        self._update_window_title()

        # v0.05: 画面サイズ（ini優先→無ければ画面から自動）
        self.ui_settings = ui_settings or {}
        self.ini_path = ini_path or str((_base_dir() / DEFAULT_INI).resolve())
        apply_initial_geometry(self, self.ui_settings)
        attach_geometry_saver(self, self.ini_path)

        self.all_cards = cards[:]
        self.filtered_cards = cards[:]
        self.cards = cards[:]
        self.index = 0
        self.topic_tag = None
        self.reverse_mode = False
        self.random_mode = False

        self.correct_count = 0
        self.wrong_count = 0
        self._checked_this_card = False
        self._last_is_ok = None  # v0.08: 直近の判定（保存に使う）

        self.lower_mode = "answer"  # 初期は正解モード
        self.lower_modes = ["answer"]
        self.q_font = ("Yu Gothic UI", 20)
        self.a_font = ("Yu Gothic UI", 18)

        self.base_cards = base_cards[:]

        # --- 上部バー ---
        self.top_bar = tk.Frame(self)
        self.top_bar.pack(fill="x", pady=(6, 0))

        self.sheet_var = tk.StringVar()
        self.sheet_names = list_question_sheets(self.excel_path)
        if self.sheet_names and str(self.source_sheet) in self.sheet_names:
            self.sheet_var.set(str(self.source_sheet))
        elif self.sheet_names:
            self.sheet_var.set(self.sheet_names[0])
        else:
            self.sheet_var.set(str(self.source_sheet))

        self.sheet_combo = ttk.Combobox(
            self.top_bar,
            textvariable=self.sheet_var,
            values=self.sheet_names if self.sheet_names else [str(self.source_sheet)],
            state="readonly",
            width=28
        )
        self.sheet_combo.pack(side="left", padx=(10, 8))

        self.mode_label = tk.Label(self.top_bar, text="通常", anchor="w")
        self.mode_label.pack(side="left", padx=(4, 12))

        self.clock_label = tk.Label(self.top_bar, text="", anchor="w")
        self.clock_label.pack(side="left", padx=(0, 12))

        self.top_info = tk.Label(self.top_bar, text="", anchor="w")
        self.top_info.pack(side="left", fill="x", expand=True)

        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        # ==================================================
        # メイン領域
        #   上段（設問）と下段（正解/解説）を同じ高さにするため grid にする
        # ==================================================
        self.left_frame = tk.Frame(self)
        self.left_frame.pack(fill="both", expand=True)

        # row 0: 設問 / row 1: 中段 / row 2: 下段
        self.left_frame.grid_rowconfigure(0, weight=1, uniform="qa")
        self.left_frame.grid_rowconfigure(1, weight=0)
        self.left_frame.grid_rowconfigure(2, weight=1, uniform="qa")
        self.left_frame.grid_rowconfigure(3, weight=0)  # help
        self.left_frame.grid_columnconfigure(0, weight=1)

        # 画像パス索引
        self._qimage_map = {}
        self._aimage_map = {}
        try:
            for c in (self.base_cards or []):
                k = (sheet_key(c.get("source_sheet")), str(c.get("question_no", "")).strip())
                self._qimage_map[k] = str(c.get("question_image_path", "") or "").strip()
                self._aimage_map[k] = str(c.get("answer_image_path", c.get("image_path", "")) or "").strip()
        except Exception:
            self._qimage_map = {}
            self._aimage_map = {}

        # ---------- 上段：設問（スクロール付き） ----------
        self.q_frame = tk.Frame(self.left_frame)
        self.q_frame.grid(row=0, column=0, sticky="nsew", padx=12, pady=(8, 6))
        self.q_frame.grid_rowconfigure(0, weight=1)
        self.q_frame.grid_columnconfigure(0, weight=1)

        self.q_scroll = tk.Scrollbar(self.q_frame, orient="vertical")
        self.q_scroll.grid(row=0, column=1, sticky="ns")

        self.q_text = tk.Text(
            self.q_frame,
            wrap="word",
            font=self.q_font,
            height=6,
            yscrollcommand=self.q_scroll.set,
        )
        self.q_text.grid(row=0, column=0, sticky="nsew")
        self.q_scroll.config(command=self.q_text.yview)
        self.q_text.configure(state="disabled")

        self.question_image_panel = ImagePanel(
            self.q_frame,
            thumb_size=220,
            zoom_max=int(self.ui_settings.get("zoom_max", 1200)),
            panel_width=420,
        )
        self.question_image_panel.grid(row=0, column=2, sticky="ns", padx=(8, 0))
        self.question_image_panel.hide()

        # ---------- 中段：絞り込み/入力/ボタン ----------
        self.mid_frame = tk.Frame(self.left_frame)
        self.mid_frame.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        self.mid_frame.grid_columnconfigure(0, weight=1)

        self.btn_frame = tk.Frame(self.mid_frame)
        self.btn_frame.grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.reverse_btn = tk.Button(self.btn_frame, text="解答⇔設問", command=self.toggle_reverse_mode)
        self.reverse_btn.pack(side="left")
        self.random_btn = tk.Button(self.btn_frame, text="ランダム出題", command=self.toggle_random_mode)
        self.random_btn.pack(side="left", padx=(8, 0))
        self.topic_btn = tk.Button(self.btn_frame, text="この論点だけ復習", command=self.filter_by_current_topic)
        self.topic_btn.pack(side="left", padx=(8, 0))
        self.clear_btn = tk.Button(self.btn_frame, text="絞り込み解除", command=self.clear_filter, state="disabled")
        self.clear_btn.pack(side="left", padx=(8, 0))

        self.answer_row = tk.Frame(self.mid_frame)
        self.answer_row.grid(row=1, column=0, sticky="w")

        self.answer_label = tk.Label(self.answer_row, text="回答（入力）", anchor="w")
        self.answer_label.pack(side="left")

        self.ox_var = tk.StringVar(value="")
        self.ox_frame = tk.Frame(self.answer_row)
        self.ox_frame.pack(side="left", padx=(10, 0))

        self.btn_o = tk.Button(
            self.ox_frame,
            text="○",
            width=3,
            font=("Yu Gothic UI", 11, "bold"),
            command=lambda: self.set_answer_symbol("○"),
            takefocus=0,
        )
        self.btn_x = tk.Button(
            self.ox_frame,
            text="×",
            width=3,
            font=("Yu Gothic UI", 11, "bold"),
            command=lambda: self.set_answer_symbol("×"),
            takefocus=0,
        )
        self.btn_o.pack(side="left", padx=(0, 4))
        self.btn_x.pack(side="left")

        self.answer_input_frame = tk.Frame(self.mid_frame)
        self.answer_input_frame.grid(row=2, column=0, sticky="ew", pady=(4, 8))
        self.answer_input_frame.grid_columnconfigure(0, weight=1)

        self.answer_text = tk.Text(self.answer_input_frame, wrap="word", font=self.a_font, height=4)
        self.answer_text.grid(row=0, column=0, sticky="ew")

        self.action_frame = tk.Frame(self.mid_frame)
        self.action_frame.grid(row=3, column=0, sticky="ew")
        # ボタンは pack のまま
        self.check_btn = tk.Button(self.action_frame, text="回答(F1)", command=self.check_answer)
        self.toggle_answer_explain_btn = tk.Button(self.action_frame, text="正解(F2)", command=self.toggle_answer_explain)
        self.save_answer_btn = tk.Button(self.action_frame, text="回答保存(F3)", command=self.save_answer_log)
        # v0.09: 自己採点（長文/表現ゆれ救済）
        self.self_ok_btn = tk.Button(self.action_frame, text="正解にする(F10)", command=lambda: self.self_grade(True))
        self.self_ng_btn = tk.Button(self.action_frame, text="不正解にする(F11)", command=lambda: self.self_grade(False))

        self.bookmark_set_btn = tk.Button(self.action_frame, text="しおり設定/解除(F5)", command=self.toggle_bookmark)
        self.bookmark_clear_btn = tk.Button(self.action_frame, text="しおり全解除(F6)", command=self.clear_all_bookmarks)
        self.prev_btn = tk.Button(self.action_frame, text="前問題(F7)", command=self.prev_card)
        self.next_btn = tk.Button(self.action_frame, text="次問題(F8)", command=self.next_card)
        self.bookmark_next_btn = tk.Button(self.action_frame, text="次のしおり(F9)", command=self.goto_next_bookmark)
        self.exit_btn = tk.Button(self.action_frame, text="終了(ESC)", command=self.destroy)

        self.check_btn.pack(side="left")
        self.toggle_answer_explain_btn.pack(side="left", padx=(8, 0))
        self.save_answer_btn.pack(side="left", padx=(8, 0))
        self.bookmark_set_btn.pack(side="left", padx=(16, 0))
        self.bookmark_clear_btn.pack(side="left", padx=(8, 0))
        self.prev_btn.pack(side="left", padx=(16, 0))
        self.next_btn.pack(side="left", padx=(8, 0))
        self.bookmark_next_btn.pack(side="left", padx=(16, 0))
        self.self_ok_btn.pack(side="left", padx=(16, 0))
        self.self_ng_btn.pack(side="left", padx=(8, 0))
        self.exit_btn.pack(side="right")

        # ---------- 下段：結果 + 正解/解説（スクロール） + 画像 ----------
        self.bottom_frame = tk.Frame(self.left_frame)
        self.bottom_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 4))
        self.bottom_frame.grid_rowconfigure(1, weight=1)
        self.bottom_frame.grid_columnconfigure(0, weight=1)
        self.bottom_frame.grid_columnconfigure(1, weight=0)

        # 結果表示（正解/不正解）＋モード表示（[解説]/[語呂合せ]）
        self.result_line = tk.Frame(self.bottom_frame)
        self.result_line.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 4))

        self.result_label = tk.Label(self.result_line, text="", anchor="w", font=("Yu Gothic UI", 14, "bold"))
        self.result_label.pack(side="left")

        # v0.08+: 下段表示モードバッジ（解説/語呂合せの時だけ表示）
        self.lower_mode_badge = tk.Label(self.result_line, text="", anchor="w", font=("Yu Gothic UI", 12, "bold"))
        self.lower_mode_badge.pack(side="left", padx=(8, 0))

        self.correct_area = tk.Frame(self.bottom_frame)
        self.correct_area.grid(row=1, column=0, sticky="nsew")

        # 左：テキスト（スクロール付き）
        self.correct_text_frame = tk.Frame(self.correct_area)
        self.correct_text_frame.pack(side="left", fill="both", expand=True)

        self.correct_scroll = tk.Scrollbar(self.correct_text_frame, orient="vertical")
        self.correct_scroll.pack(side="right", fill="y")

        self.correct_text = tk.Text(
            self.correct_text_frame,
            wrap="word",
            font=self.a_font,
            height=6,
            yscrollcommand=self.correct_scroll.set,
        )
        self.correct_text.pack(side="left", fill="both", expand=True)
        self.correct_scroll.config(command=self.correct_text.yview)
        self.correct_text.configure(state="disabled")

        # 右：画像（回答表示時のみ表示）
        self.answer_image_panel = ImagePanel(
            self.correct_area,
            thumb_size=240,
            zoom_max=int(self.ui_settings.get("zoom_max", 1200)),
            panel_width=420,
        )
        self.answer_image_panel.hide()

        self.help = tk.Label(
            self.left_frame,
            text="判定/保存はボタンのみ。※保存はF1判定後。『回答保存』で回答シートへ登録（削除はExcel手動）",
            anchor="w",
            padx=10,
            justify="left",
            wraplength=980
        )
        self.help.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 8))

        # ==================================================
        # しおり（FlashCardsShiori.dat）
        # ==================================================
        self.shiori_path = self._get_shiori_path()
        self.shiori_data = self._load_shiori_data()

        # ==================================================
        # Key bindings
        # ==================================================
        def _safe_invoke(btn: tk.Button):
            try:
                if str(btn.cget("state")) != "disabled":
                    btn.invoke()
            except Exception:
                pass

        self.bind_all("<Prior>",   lambda e: _safe_invoke(self.prev_btn))  # PageUp
        self.bind_all("<Next>",    lambda e: _safe_invoke(self.next_btn))  # PageDown
        self.bind_all("<Page_Up>",   lambda e: _safe_invoke(self.prev_btn))
        self.bind_all("<Page_Down>", lambda e: _safe_invoke(self.next_btn))

        self.bind_all("<F1>", lambda e: _safe_invoke(self.check_btn))
        self.bind_all("<F2>", lambda e: _safe_invoke(self.toggle_answer_explain_btn))
        self.bind_all("<F3>", lambda e: _safe_invoke(self.save_answer_btn))
        self.bind_all("<F5>", lambda e: _safe_invoke(self.bookmark_set_btn))
        self.bind_all("<F6>", lambda e: _safe_invoke(self.bookmark_clear_btn))
        self.bind_all("<F7>", lambda e: _safe_invoke(self.prev_btn))
        self.bind_all("<F8>", lambda e: _safe_invoke(self.next_btn))
        self.bind_all("<F9>", lambda e: self._safe_goto_next_bookmark())
        # v0.09: 自己採点
        self.bind_all("<F10>", lambda e: self.self_grade(True))
        self.bind_all("<F11>", lambda e: self.self_grade(False))
        self.bind_all("<Escape>", lambda e: _safe_invoke(self.exit_btn))

        self.random_mode = bool(initial_random_mode)
        self._rebuild_cards_view(reset_index=True)
        self._tick_clock()
        self.render()

    # ---------------- UI helpers ----------------
    def set_text(self, widget: tk.Text, value: str):
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", value)
        widget.configure(state="disabled")

    def _get_display_question(self, item: dict) -> str:
        if self.reverse_mode:
            return f"{item.get('answer','')}"
        return f"{item.get('question','')}"

    def _get_display_answer(self, item: dict) -> str:
        if self.reverse_mode:
            text = (item.get("question", "") or "").strip()
            if not text:
                text = "（この問題は設問セルが空です）"
            return text
        text = (item.get("answer", "") or "").strip()
        if not text:
            text = "（この問題は解答セルが空です）"
        return text

    def _build_lower_modes(self, item: dict) -> list:
        modes = ["answer"]
        if has_trim_value(item.get("explanation", "")):
            modes.append("explain")
        if has_trim_value(item.get("mnemonic", "")):
            modes.append("mnemonic")
        return modes

    def _build_f2_button_label(self, item: dict) -> str:
        labels = ["設問" if self.reverse_mode else "正解"]
        if has_trim_value(item.get("explanation", "")):
            labels.append("解説")
        if has_trim_value(item.get("mnemonic", "")):
            labels.append("語呂合せ")
        return "/".join(labels) + "(F2)"

    def _apply_current_f2_label(self) -> None:
        try:
            item = self.current()
            self.lower_modes = self._build_lower_modes(item)
            self.lower_mode = "answer"
            if hasattr(self, "toggle_answer_explain_btn"):
                self.toggle_answer_explain_btn.configure(text=self._build_f2_button_label(item))
        except Exception:
            pass
    def _current_card_key(self, item: dict | None) -> tuple[str, str]:
        if not item:
            return ("", "")
        return (
            str(item.get("source_sheet", self.source_sheet) or "").strip(),
            str(item.get("question_no", "") or "").strip(),
        )

    def _rebuild_cards_view(self, *, keep_current: bool = False, reset_index: bool = False) -> None:
        source_cards = (getattr(self, "filtered_cards", None) or self.all_cards or [])[:]
        current_key = None
        if keep_current and getattr(self, "cards", None):
            try:
                current_key = self._current_card_key(self.current())
            except Exception:
                current_key = None

        self.cards = source_cards[:]
        if self.random_mode:
            random.shuffle(self.cards)

        if not self.cards:
            self.index = 0
            return

        if reset_index:
            self.index = 0
            return

        if current_key:
            for i, card in enumerate(self.cards):
                if self._current_card_key(card) == current_key:
                    self.index = i
                    return

        if self.index >= len(self.cards):
            self.index = len(self.cards) - 1
        if self.index < 0:
            self.index = 0

    def _update_random_button(self) -> None:
        try:
            if hasattr(self, "random_btn"):
                # トグル式らしく、ボタンには「次に切り替わる先」を表示する
                self.random_btn.configure(text=("通常出題" if self.random_mode else "ランダム出題"))
        except Exception:
            pass


    def _tick_clock(self):
        try:
            self.clock_label.configure(text=time.strftime("%Y/%m/%d %H:%M:%S"))
        except Exception:
            pass
        self.after(1000, self._tick_clock)

    def _update_lower_mode_badge(self) -> None:
        """下段が『解説』『語呂合せ』表示のときだけバッジを表示する（未判定時は非表示）。"""
        try:
            if not getattr(self, "_checked_this_card", False):
                txt = ""
            else:
                mode = getattr(self, "lower_mode", "answer")
                if mode == "explain":
                    txt = "[解説]"
                elif mode == "mnemonic":
                    txt = "[語呂合せ]"
                else:
                    txt = ""
            if hasattr(self, "lower_mode_badge"):
                self.lower_mode_badge.configure(text=txt)
        except Exception:
            pass

    def current(self):
        return self.cards[self.index]

    def set_answer_symbol(self, symbol: str):
        try:
            self.answer_text.delete("1.0", "end")
            self.answer_text.insert("1.0", symbol)
            if hasattr(self, "ox_var"):
                self.ox_var.set(symbol)
            self.answer_text.focus_set()
        except Exception:
            pass

    def clear_answer_area(self):
        self._apply_current_f2_label()
        self.answer_text.delete("1.0", "end")
        if hasattr(self, "ox_var"):
            self.ox_var.set("")
        self.result_label.configure(text="")
        self.set_text(self.correct_text, "")
        self._hide_answer_image()
        self._checked_this_card = False
        self._last_is_ok = None
        self._update_lower_mode_badge()
        self.answer_text.focus_set()

    def update_nav_buttons(self):
        self.prev_btn.configure(state=("disabled" if self.index <= 0 else "normal"))
        self.next_btn.configure(state=("disabled" if self.index >= len(self.cards) - 1 else "normal"))

    def _update_mode_label(self):
        try:
            # セレクトボックス横には、現在の出題順モードだけをわかりやすく表示する
            self.mode_label.configure(text=("ランダム出題" if self.random_mode else "通常出題"))
            if hasattr(self, "reverse_btn"):
                self.reverse_btn.configure(text=("設問⇔解答" if self.reverse_mode else "解答⇔設問"))
            self._update_random_button()
        except Exception:
            pass

    def _read_sheet_log_map(self) -> dict:
        records = {}
        try:
            log_path = self._get_progress_log_path()
            if not os.path.exists(log_path):
                return records
            with open(log_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.rstrip("\n")
                    if not line.strip():
                        continue
                    parts = line.split("	")
                    while len(parts) < 3:
                        parts.append("")
                    dt, qno, result = parts[0], parts[1].strip(), parts[2].strip().upper()
                    if qno:
                        records[qno] = (dt, result)
        except Exception:
            return {}
        return records

    def _get_progress_stats(self):
        try:
            total = len(self.base_cards) if getattr(self, 'base_cards', None) else len(self.cards)
            if total <= 0:
                return 0, 0, 0.0, 0.0
            log_map = self._read_sheet_log_map()
            answered = 0
            ok_count = 0
            valid_qnos = {str(c.get('question_no', '')).strip() for c in (self.base_cards or self.cards) if str(c.get('question_no', '')).strip()}
            for qno, (_dt, result) in log_map.items():
                if qno in valid_qnos:
                    answered += 1
                    if result == 'OK':
                        ok_count += 1
            answered_rate = (answered / total * 100.0) if total > 0 else 0.0
            ok_rate = (ok_count / total * 100.0) if total > 0 else 0.0
            return answered, ok_count, answered_rate, ok_rate
        except Exception:
            return 0, 0, 0.0, 0.0

    def update_top_info(self):
        total = len(self.cards)
        page = self.index + 1
        ok = self.correct_count
        ng = self.wrong_count
        tried = ok + ng
        rate = (ok / tried * 100.0) if tried > 0 else 0.0
        prog_total = len(self.base_cards) if getattr(self, 'base_cards', None) else total
        answered_count, progress_ok_count, answered_rate, progress_ok_rate = self._get_progress_stats()

        no_val = self.current().get('question_no', '')
        prefix = ""
        if self.is_current_bookmarked():
            prefix = "★しおり   "

        self.top_info.configure(
            text=(
                f"{prefix}No:{no_val}   {page} / {total}   ○:{ok}   ✕:{ng}   正答率:{rate:.1f}%"
                f"   回答進捗:{answered_count}/{prog_total} ({answered_rate:.1f}%)"
                f"   正解進捗:{progress_ok_count}/{prog_total} ({progress_ok_rate:.1f}%)"
            )
        )

    def render(self):
        item = self.current()
        self._set_current_position(save=True)
        self.set_text(self.q_text, self._get_display_question(item))
        self._show_question_image(item)
        self.update_nav_buttons()
        self._update_mode_label()
        self.update_top_info()
        self._update_lower_mode_badge()
        self.update_bookmark_ui()
        self.clear_answer_area()

    # ---------------- Sheet change ----------------
    def _update_window_title(self) -> None:
        """iniの app_title / app_version を反映したウィンドウタイトル。"""
        base = self.app_title or "暗記カード"
        ver = self.app_version.strip()
        head = f"{base} {ver}".strip() if ver else base
        try:
            self.title(f"{head} - {Path(self.excel_path).name} [{self.source_sheet}]")
        except Exception:
            # タイトル更新失敗でも落とさない
            pass

    def on_sheet_selected(self, event=None):
        sheet_name = self.sheet_var.get().strip()
        if not sheet_name:
            return
        # wrong_sheet は除外しない（ユーザが選ぶ自由。ただしカード0件ならエラー）
        try:
            base_cards = load_cards(self.excel_path, sheet_name, data_start_row=self.data_start_row_default)
            if not base_cards:
                raise ValueError("カードが0件です。")

            for c in base_cards:
                c["source_sheet"] = str(sheet_name)

            self.base_cards = base_cards[:]
            self.all_cards = base_cards[:]
            self.filtered_cards = base_cards[:]
            self.cards = base_cards[:]
            self.index = 0
            self.topic_tag = None
            self.correct_count = 0
            self.wrong_count = 0
            self._checked_this_card = False
            self._last_is_ok = None
            self.source_sheet = str(sheet_name)

            self._update_window_title()

            if hasattr(self, "clear_btn"):
                self.clear_btn.configure(state="disabled")

            self._rebuild_cards_view(reset_index=True)

            # 画像索引更新
            self._qimage_map = {}
            self._aimage_map = {}
            try:
                for c in (self.base_cards or []):
                    k = (sheet_key(c.get("source_sheet")), str(c.get("question_no", "")).strip())
                    self._qimage_map[k] = str(c.get("question_image_path", "") or "").strip()
                    self._aimage_map[k] = str(c.get("answer_image_path", c.get("image_path", "")) or "").strip()
            except Exception:
                self._qimage_map = {}
                self._aimage_map = {}

            self.render()
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    # ---------------- Image helpers ----------------
    def _get_question_image_path(self, item: dict) -> str:
        try:
            if self.reverse_mode:
                img_path = str(item.get("answer_image_path", item.get("image_path", "")) or "").strip()
                if not img_path:
                    k = (sheet_key(item.get("source_sheet")), str(item.get("question_no", "")).strip())
                    img_path = str(self._aimage_map.get(k, "") or "").strip()
            else:
                img_path = str(item.get("question_image_path", "") or "").strip()
                if not img_path:
                    k = (sheet_key(item.get("source_sheet")), str(item.get("question_no", "")).strip())
                    img_path = str(self._qimage_map.get(k, "") or "").strip()
            return img_path
        except Exception:
            return ""

    def _get_answer_image_path(self, item: dict) -> str:
        try:
            if self.reverse_mode:
                img_path = str(item.get("question_image_path", "") or "").strip()
                if not img_path:
                    k = (sheet_key(item.get("source_sheet")), str(item.get("question_no", "")).strip())
                    img_path = str(self._qimage_map.get(k, "") or "").strip()
            else:
                img_path = str(item.get("answer_image_path", item.get("image_path", "")) or "").strip()
                if not img_path:
                    k = (sheet_key(item.get("source_sheet")), str(item.get("question_no", "")).strip())
                    img_path = str(self._aimage_map.get(k, "") or "").strip()
            return img_path
        except Exception:
            return ""

    def _hide_question_image(self) -> None:
        try:
            self.question_image_panel.set_image("")
        except Exception:
            pass
        try:
            self.question_image_panel.hide()
        except Exception:
            pass

    def _show_question_image(self, item: dict) -> None:
        try:
            img_path = self._get_question_image_path(item)
            if img_path:
                self.question_image_panel.set_image(img_path)
                self.question_image_panel.show()
            else:
                self._hide_question_image()
        except Exception:
            self._hide_question_image()

    def _hide_answer_image(self) -> None:
        try:
            self.answer_image_panel.set_image("")
        except Exception:
            pass
        try:
            self.answer_image_panel.hide()
        except Exception:
            pass

    def _show_answer_image(self, item: dict) -> None:
        try:
            img_path = self._get_answer_image_path(item)
            if img_path:
                self.answer_image_panel.set_image(img_path)
                self.answer_image_panel.show()
            else:
                self._hide_answer_image()
        except Exception:
            self._hide_answer_image()

    # ---------------- Answer / Explain ----------------

    # ==================================================
    # v0.09: 自己採点（長文/表現ゆれ救済）
    #   - F10: 正解にする
    #   - F11: 不正解にする
    #   - まだ判定していない場合のみ、正解数/不正解数を加算
    # ==================================================
    def self_grade(self, is_ok: bool) -> None:
        try:
            item = self.current()

            self._last_is_ok = bool(is_ok)
            if not getattr(self, "_checked_this_card", False):
                if is_ok:
                    self.correct_count += 1
                else:
                    self.wrong_count += 1
                self._checked_this_card = True
                self.update_top_info()

            if is_ok:
                self.result_label.configure(text="✅ 正解！（自己採点）")
            else:
                self.result_label.configure(text="✖ 不正解（自己採点）")

            # 下段（正解/解説）を更新し、画像も表示
            self._refresh_lower_text()
            self._show_answer_image(item)
        except Exception:
            # 自己採点で落ちない（ループ地獄回避）
            pass


    def toggle_answer_explain(self):
        modes = getattr(self, "lower_modes", None) or self._build_lower_modes(self.current())
        try:
            idx = modes.index(getattr(self, "lower_mode", "answer"))
        except ValueError:
            idx = 0
        self.lower_mode = modes[(idx + 1) % len(modes)]
        self._update_lower_mode_badge()
        if getattr(self, "_checked_this_card", False):
            self._refresh_lower_text()

    def _refresh_lower_text(self):
        item = self.current()
        if self.lower_mode == "answer":
            text = self._get_display_answer(item)
        elif self.lower_mode == "explain":
            text = (item.get("explanation", "") or "").strip()
            if not text:
                text = "（解説なし）"
        else:
            text = (item.get("mnemonic", "") or "").strip()
            if not text:
                text = "（語呂合せなし）"
        self.set_text(self.correct_text, text)

    def check_answer(self):
        item = self.current()
        user_raw = self.answer_text.get("1.0", "end-1c").strip()
        correct_raw = self._get_display_answer(item).strip()
        first_check = not self._checked_this_card

        # 逆出題時は元の設問、通常時は元の解答が比較対象
        if not correct_raw or correct_raw in ("（この問題は解答セルが空です）", "（この問題は設問セルが空です）"):
            self.result_label.configure(text="ℹ️ 正解（比較対象セルが空です）")
            self._last_is_ok = True

            if first_check:
                self.correct_count += 1
                self._checked_this_card = True
                self.update_top_info()
                self.upsert_progress_log(True)

            self._refresh_lower_text()
            self._show_answer_image(item)
            self._update_lower_mode_badge()
            return

        strict_ok = (user_raw.strip() == correct_raw.strip())
        lenient_ok = (normalize_answer(user_raw) == normalize_answer(correct_raw))
        is_ok = strict_ok or lenient_ok

        # 初回判定のみカウント＋log書き込み
        if first_check:
            if is_ok:
                self.correct_count += 1
            else:
                self.wrong_count += 1
            self._checked_this_card = True
            self.update_top_info()
            self.upsert_progress_log(is_ok)

        self._last_is_ok = bool(is_ok)

        if is_ok:
            self.result_label.configure(text="✅ 正解！")
        else:
            self.result_label.configure(text="✖ 不正解")

        self._refresh_lower_text()
        self._show_answer_image(item)
        self._update_lower_mode_badge()

    def save_answer_log(self):
        """F3: 回答保存（F1判定後のみ）"""
        if not self._checked_this_card or self._last_is_ok is None:
            messagebox.showinfo("情報", "先に「回答(F1)」で判定してください（保存は判定後のみ）。")
            return

        item = self.current()
        try:
            upsert_answer_log(
                self.excel_path,
                self.wrong_sheet,
                source_sheet=item.get("source_sheet", self.source_sheet),
                question_no=item.get("question_no", ""),
                q=item.get("question", ""),
                a=item.get("answer", ""),
                question_image_path=item.get("question_image_path", ""),
                answer_image_path=item.get("answer_image_path", item.get("image_path", "")),
                image_path=item.get("answer_image_path", item.get("image_path", "")),
                explanation=item.get("explanation", ""),
                subject=item.get("subject", ""),
                tags_list=item.get("tags", []),
                keywords_list=item.get("keywords", []),
                is_ok=bool(self._last_is_ok),
                wrong_start_row=self.wrong_start_row,
            )
            qno = item.get("question_no", "")
            if self._last_is_ok:
                self.result_label.configure(text=f"💾 正解保存：No.{qno}")
            else:
                self.result_label.configure(text=f"💾 不正解保存：No.{qno}")
            self.after(900, lambda: self.result_label.configure(text=""))
        except Exception as e:
            messagebox.showerror("保存エラー", str(e))

    def toggle_reverse_mode(self):
        try:
            self.reverse_mode = not self.reverse_mode
            self.lower_mode = "answer"
            self._checked_this_card = False
            self._last_is_ok = None
            self.render()
        except Exception:
            pass

    def toggle_random_mode(self):
        try:
            self.random_mode = not self.random_mode
            self._rebuild_cards_view(keep_current=True)
            self.render()
        except Exception:
            pass

    # ---------------- Filter by tag ----------------
    def filter_by_current_topic(self):
        item = self.current()
        tags = item.get("tags", []) or []
        if not tags:
            return
        tag = tags[0]
        filtered = [c for c in self.all_cards if tag in (c.get("tags", []) or [])]
        if not filtered:
            self.filtered_cards = self.all_cards[:]
            self.topic_tag = None
            messagebox.showinfo("情報", "この論点タグに該当する問題が見つかりませんでした。")
            return
        self.topic_tag = tag
        self.filtered_cards = filtered
        self.index = 0
        self.clear_btn.configure(state="normal")
        self._rebuild_cards_view(reset_index=True)
        self.render()

    def clear_filter(self):
        self.filtered_cards = self.all_cards[:]
        self.topic_tag = None
        self.index = 0
        self.clear_btn.configure(state="disabled")
        self._rebuild_cards_view(reset_index=True)
        self.render()

    # ---------------- Navigation ----------------
    def next_card(self):
        if self.index < len(self.cards) - 1:
            self.index += 1
            self.render()

    def prev_card(self):
        if self.index > 0:
            self.index -= 1
            self.render()

    # ==================================================
    # しおり (FlashCardsShiori.dat)
    # ==================================================
    @staticmethod
    def _get_app_dir() -> str:
        try:
            if getattr(sys, "frozen", False):
                return os.path.dirname(sys.executable)
        except Exception:
            pass
        return os.path.dirname(os.path.abspath(__file__))

    def _get_shiori_path(self) -> str:
        return os.path.join(self._get_app_dir(), "FlashCardsShiori.dat")

    def _normalize_excel_path(self) -> str:
        try:
            return os.path.abspath(self.excel_path)
        except Exception:
            return str(self.excel_path)

    def _shiori_key(self) -> str:
        return f"{self._normalize_excel_path()}||{self.source_sheet}"

    def _load_shiori_data(self) -> dict:
        data = {
            "current": None,
            "bookmarks": {},
        }
        try:
            if not os.path.exists(self.shiori_path):
                return data
            with open(self.shiori_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if not row:
                        continue
                    rtype = (row[0] or "").strip().upper()
                    if rtype not in ("CURRENT", "BOOKMARK"):
                        continue
                    if len(row) < 4:
                        continue
                    excel_path = (row[1] or "").strip()
                    sheet_name = (row[2] or "").strip()
                    try:
                        idx = int(str(row[3]).strip())
                    except Exception:
                        continue
                    key = f"{excel_path}||{sheet_name}"
                    if rtype == "CURRENT":
                        data["current"] = {"excel_path": excel_path, "sheet_name": sheet_name, "index": idx}
                    else:
                        data["bookmarks"].setdefault(key, []).append(idx)

            for k, v in list(data["bookmarks"].items()):
                uniq = sorted({int(x) for x in v if isinstance(x, int) or str(x).isdigit()})
                data["bookmarks"][k] = uniq

        except Exception:
            return {"current": None, "bookmarks": {}}

        return data

    def _save_shiori_data(self) -> None:
        try:
            current = self.shiori_data.get("current")
            bookmarks = self.shiori_data.get("bookmarks", {})

            with open(self.shiori_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                if not current:
                    current = {
                        "excel_path": self._normalize_excel_path(),
                        "sheet_name": self.source_sheet,
                        "index": int(self.index),
                    }
                writer.writerow(["CURRENT", current["excel_path"], current["sheet_name"], int(current["index"])])

                for key, idx_list in bookmarks.items():
                    try:
                        excel_path, sheet_name = key.split("||", 1)
                    except Exception:
                        continue
                    for idx in sorted({int(i) for i in idx_list}):
                        writer.writerow(["BOOKMARK", excel_path, sheet_name, idx])

        except Exception:
            pass

    def _set_current_position(self, save: bool = True) -> None:
        try:
            self.shiori_data["current"] = {
                "excel_path": self._normalize_excel_path(),
                "sheet_name": self.source_sheet,
                "index": int(self.index),
            }
            if save:
                self._save_shiori_data()
        except Exception:
            pass

    def _current_bookmark_list(self):
        key = self._shiori_key()
        try:
            lst = self.shiori_data.get("bookmarks", {}).get(key, [])
            max_idx = len(self.cards) - 1
            cleaned = sorted({int(i) for i in lst if 0 <= int(i) <= max_idx})
            self.shiori_data.setdefault("bookmarks", {})[key] = cleaned
            return cleaned
        except Exception:
            return []

    def is_current_bookmarked(self) -> bool:
        try:
            return int(self.index) in set(self._current_bookmark_list())
        except Exception:
            return False

    def update_bookmark_ui(self) -> None:
        try:
            has_any = len(self._current_bookmark_list()) > 0
            self.bookmark_set_btn.configure(state="normal")
            self.bookmark_clear_btn.configure(state=("normal" if has_any else "disabled"))
        except Exception:
            pass

    def toggle_bookmark(self) -> None:
        try:
            key = self._shiori_key()
            bm = self.shiori_data.setdefault("bookmarks", {}).setdefault(key, [])
            cur = int(self.index)
            s = set(int(i) for i in bm)
            if cur in s:
                s.remove(cur)
            else:
                s.add(cur)
            self.shiori_data["bookmarks"][key] = sorted(s)
            self._set_current_position(save=False)
            self._save_shiori_data()
        except Exception:
            pass
        self.update_top_info()
        self._update_lower_mode_badge()
        self.update_bookmark_ui()

    def clear_all_bookmarks(self) -> None:
        try:
            key = self._shiori_key()
            if key in self.shiori_data.get("bookmarks", {}):
                self.shiori_data["bookmarks"][key] = []
            self._set_current_position(save=False)
            self._save_shiori_data()
        except Exception:
            pass
        self.update_top_info()
        self._update_lower_mode_badge()
        self.update_bookmark_ui()

    def goto_next_bookmark(self) -> None:
        try:
            bms = self._current_bookmark_list()
            if not bms:
                return
            cur = int(self.index)
            nxt = None
            for i in bms:
                if i > cur:
                    nxt = i
                    break
            if nxt is None:
                nxt = bms[0]

            self.index = int(nxt)
            self._checked_this_card = False
            self._last_is_ok = None
            self._set_current_position(save=True)
            self.render()
        except Exception:
            pass


    def _safe_sheet_filename(self, name: str) -> str:
        s = str(name or "").strip()
        if not s:
            s = "unknown"
        for ch in r'\\/:*?"<>|':
            s = s.replace(ch, "_")
        return s

    def _get_progress_log_path(self) -> str:
        log_dir = os.path.join(self._get_app_dir(), "logs")
        os.makedirs(log_dir, exist_ok=True)
        return os.path.join(log_dir, f"{self._safe_sheet_filename(self.source_sheet)}.log")

    def upsert_progress_log(self, is_ok: bool) -> None:
        """F1判定時の最新結果を、設問Noキーで上書き保存する。"""
        try:
            item = self.current()
            qno = str(item.get("question_no", "")).strip()
            if not qno:
                return

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            result = "OK" if is_ok else "NG"
            log_path = self._get_progress_log_path()

            rows = []
            found = False

            if os.path.exists(log_path):
                with open(log_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.rstrip("\r\n")
                        if not line.strip():
                            continue
                        parts = line.split("\t")
                        while len(parts) < 3:
                            parts.append("")
                        old_dt, old_qno, old_result = parts[0], parts[1], parts[2]

                        if old_qno == qno:
                            rows.append([now, qno, result])
                            found = True
                        else:
                            rows.append([old_dt, old_qno, old_result])

            if not found:
                rows.append([now, qno, result])

            def sort_key(r):
                try:
                    return (0, int(str(r[1]).strip()))
                except Exception:
                    return (1, str(r[1]).strip())

            rows.sort(key=sort_key)

            with open(log_path, "w", encoding="utf-8", newline="") as f:
                for dt_str, no_str, rs in rows:
                    f.write(f"{dt_str}\t{no_str}\t{rs}\n")
        except Exception:
            # log書き込み失敗で本体を落とさない
            pass


    def _safe_goto_next_bookmark(self):
        try:
            self.goto_next_bookmark()
        except Exception:
            pass


def parse_args():
    p = argparse.ArgumentParser(description="Excelから暗記カードGUIを起動します（安定版）")
    p.add_argument("--excel", default=None, help="Excelファイルパス（指定時のみiniより優先）")
    p.add_argument("--sheet", default=None, help="シート名 or 番号（0始まり）。未指定は先頭シート")
    p.add_argument("--random", action="store_true", help="ランダム出題（通常モード）")
    p.add_argument("--wrong-sheet", default=None, help="回答シート名（指定時のみiniより優先）")
    p.add_argument("--reverse", action="store_true", help="設問/解答を逆転して開始")
    return p.parse_args()


def main():
    settings = load_settings()
    args = parse_args()

    excel_path = args.excel if args.excel else settings["EXCEL_PATH"]
    wrong_sheet = args.wrong_sheet if args.wrong_sheet else settings["wrong_sheet"]

    sheet_arg = args.sheet if args.sheet is not None else settings.get("initial_sheet", "sheet0")
    sheet_name = resolve_sheet_name(excel_path, sheet_arg)

    try:
        base_cards = load_cards(
            excel_path,
            sheet_name,
            data_start_row=settings.get("data_start_row_default", DATA_START_ROW_DEFAULT),
        )
        if not base_cards:
            raise ValueError("カードが0件です。問題列に1行以上入力してください。")

        for c in base_cards:
            c["source_sheet"] = str(sheet_name)

        cards = base_cards[:]

    except Exception as e:
        tk.Tk().withdraw()
        messagebox.showerror("エラー", str(e))
        return

    app = FlashcardsApp(
        cards,
        excel_path=excel_path,
        source_sheet=str(sheet_name),
        wrong_sheet=str(wrong_sheet),
        base_cards=base_cards,
        app_title=settings.get("app_title", "暗記カード"),
        app_version=APP_VERSION,
        ui_settings=settings.get("ui", {}),
        ini_path=settings.get("ini_path"),
        data_start_row_default=settings.get("data_start_row_default", DATA_START_ROW_DEFAULT),
        wrong_start_row=settings.get("wrong_start_row", WRONG_START_ROW),
        initial_random_mode=bool(args.random),
    )
    if args.reverse:
        try:
            app.reverse_mode = True
            app.render()
        except Exception:
            pass
    app.mainloop()


if __name__ == "__main__":
    main()